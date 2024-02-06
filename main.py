import os
from dotenv import load_dotenv
from langchain.llms import AI21
from langchain.prompts import PromptTemplate
from langchain.chains import LLMChain
from openpyxl import load_workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

load_dotenv()

intention = input("what's the emails intention?")
template = """Create me an email with the Intention: {intention}
            the details of the recipient for the context of the mail is {details}
            Create me an email for this recipient and mention in the email how the intention I mentioned matches the details of the recipient like his professions or business that I mentioned in the details"""
prompt = PromptTemplate(template=template, input_variables=["intention", "details"])

llm = AI21(ai21_api_key=os.getenv("ai21_token"))
llm_chain = LLMChain(prompt=prompt, llm=llm)

excel_file_path = "./contacts.xlsx"

wb = load_workbook(excel_file_path)
sheet = wb.active
row_counter = 2 

for row in sheet.iter_rows(min_row=row_counter, values_only=True):
    print("-----------------",row)
    name, email, phone, details = row
    print(f"Details for {name}: {details}")
    message = llm_chain.run(intention=intention, details="recipient email is {email} and name of recipient: {name} other details of recipient to use in the mail content:{details}")
    message = message.strip() 
    print(f"Name: {name}, Email: {email}, Message: {message}")
    
    msg = MIMEMultipart()
    msg['From'] = os.getenv("sender_email")
    msg['To'] = email
    msg['Subject'] = "Your AI Generated Message"

    msg.attach(MIMEText(message, 'plain'))

    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(os.getenv("sender_email"), os.getenv("email_password"))
        server.sendmail(os.getenv("sender_email"), email, msg.as_string())
        server.close()
        print(f"Email sent to {email}")
    except Exception as e:
        print(f"Failed to send email to {email}. Error: {str(e)}")

wb.save(excel_file_path)
print("Emails sent.")